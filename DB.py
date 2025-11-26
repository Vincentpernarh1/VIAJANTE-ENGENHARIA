import pandas as pd
from tkinter import *
from tkinter import ttk
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from math import ceil
import re
from PIL import Image, ImageTk
import traceback
import os
import numpy as np
import warnings 
import contextlib

# Suppress xlrd / Excel warnings
warnings.simplefilter("ignore")

# Optionally suppress all print output from the engine
with contextlib.redirect_stdout(None), contextlib.redirect_stderr(None):
    df = pd.read_excel("Template.xlsx", engine="openpyxl")
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    message=".*OLE2.*"
)
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    message=".*CODEPAGE.*"
)
warnings.filterwarnings(
    "ignore",
    message="^WARNING .*" # Hides the file size warnings which don't have a category
)



caminho_base = os.getcwd()
    

def Processar_Demandas(cod_destino, pasta_demandas="Demandas"):
    """
    Processa arquivos de demanda de uma pasta, tratando arquivos de texto/CSV
    e Excel de forma diferente, e os consolida em um único DataFrame.
    """
    # Define o caminho completo para a pasta de demandas
    caminho_pasta = os.path.join(caminho_base, pasta_demandas)

    # Verifica se a pasta de demandas existe
    if not os.path.isdir(caminho_pasta):
        print(f"Aviso: A pasta '{caminho_pasta}' não foi encontrada.")
        return pd.DataFrame()

    # Lista para armazenar os DataFrames de cada arquivo processado
    lista_dfs = []

    # Percorre todos os arquivos na pasta de demandas
    for nome_arquivo in os.listdir(caminho_pasta):
        caminho_completo_arquivo = os.path.join(caminho_pasta, nome_arquivo)
        nome_arquivo_lower = nome_arquivo.lower()
        
        try:
            # --- MANTÉM A LÓGICA ORIGINAL PARA ARQUIVOS .TXT E .CSV ---
            if nome_arquivo_lower.endswith((".txt", ".csv")):
                dados_arquivo_atual = []
                with open(caminho_completo_arquivo, "r", encoding="utf-8", errors="ignore") as arquivo:
                    linhas_a_processar = arquivo.readlines()

                # Processa cada linha extraída do arquivo de texto
                for linha in linhas_a_processar:
                    if "AUTOMATIC" in linha:
                        continue

                    linha = linha.strip()

                    # A lógica de fatiamento requer um comprimento mínimo
                    if len(linha) >= 20:
                        try:
                            # Extrai os dados com base na posição dos caracteres
                            desenho = linha[3:14]
                            cod_fornecedor = linha[-20:-11]
                            quantidade = linha[-11:].replace("+", "")

                            # Adiciona os dados extraídos à lista deste arquivo
                            dados_arquivo_atual.append({
                                "DESENHO": int(desenho.strip()),
                                "COD FORNECEDOR": int(cod_fornecedor.strip()),
                                "QTDE": int(quantidade.strip()),
                            })
                        except (ValueError, IndexError):
                            # Ignora linhas que não seguem o formato esperado
                            continue
                
                # Se dados foram extraídos do arquivo, cria um DataFrame
                if dados_arquivo_atual:
                    df_temp = pd.DataFrame(dados_arquivo_atual)
                    df_temp["COD DESTINO"] = cod_destino
                    lista_dfs.append(df_temp)

            # --- NOVA LÓGICA PARA PROCESSAR ARQUIVOS EXCEL (.XLS, .XLSX) ---
            elif nome_arquivo_lower.endswith((".xls", ".xlsx")):
                
                # Mapeamento dos nomes de coluna do arquivo Excel para os nomes desejados
                colunas_mapeamento = {
                    'DESENHO': 'DESENHO',
                    'COD ORIGEM': 'COD FORNECEDOR',
                    'ENTREGA SOLICITADA': 'QTDE',
                    'COD DESTINO': 'COD DESTINO'
                }
                
                # Lê o arquivo Excel
                df_excel = pd.read_excel(caminho_completo_arquivo)

                # Pega a lista de colunas que precisamos do arquivo original
                colunas_originais_necessarias = list(colunas_mapeamento.keys())

                # Verifica se todas as colunas necessárias existem no arquivo
                if not all(coluna in df_excel.columns for coluna in colunas_originais_necessarias):
                    print(f"Aviso: O arquivo '{nome_arquivo}' não contém todas as colunas necessárias e será ignorado.")
                    continue

                # 1. Seleciona apenas as colunas que nos interessam
                df_temp = df_excel[colunas_originais_necessarias].copy()
                
                # 2. Renomeia as colunas para o padrão final
                df_temp.rename(columns=colunas_mapeamento, inplace=True)

                # 3. Adiciona o DataFrame processado à lista para concatenação posterior
                lista_dfs.append(df_temp)

        except Exception as e:
            print(f"Erro ao processar o arquivo '{nome_arquivo}': {e}")
            continue

    # --- LÓGICA FINAL PARA CONSOLIDAR OS DADOS ---
    # Se a lista de DataFrames estiver vazia, retorna um DataFrame vazio
    if not lista_dfs:
        print("Nenhum dado válido foi processado.")
        return pd.DataFrame()
    
    # Concatena todos os DataFrames da lista em um único DataFrame final
    df_final = pd.concat(lista_dfs, ignore_index=True)

    # Garante que as colunas numéricas tenham o tipo de dados correto, tratando possíveis erros
    # Esta etapa é importante para garantir a consistência entre os dados de texto e Excel
    colunas_numericas = ["DESENHO", "COD FORNECEDOR", "QTDE"]
    for col in colunas_numericas:
        df_final[col] = pd.to_numeric(df_final[col], errors='coerce')

    # Remove linhas onde a conversão numérica falhou (resultando em NaT/NaN)
    df_final.dropna(subset=colunas_numericas, inplace=True)

    # Converte colunas para inteiro após remover os nulos
    for col in colunas_numericas:
        df_final[col] = df_final[col].astype(int)
    
    return df_final

# Exemplo de como chamar a função
# df_processado = Processar_Demandas(cod_destino="BR01")
# print(df_processado)







def desenhar_caminhoes(canvas, ocupacao, caminhao_img):
    canvas.delete("all")

    if caminhao_img is None:
        return

    quad_por_caminhao = 35
    quad_linha = 7
    quad_coluna = 5
    quad_largura = 14
    quad_altura = 14
    margem = 10

    total_quads = ceil(ocupacao * quad_por_caminhao / 100)
    max_caminhoes = 3
    num_caminhoes = min((total_quads - 1) // quad_por_caminhao + 1, max_caminhoes)

    for caminhao_idx in range(num_caminhoes):
        # Posição em "grade" 2 acima, 1 abaixo
        if caminhao_idx < 2:
            x_offset = margem + caminhao_idx * 180  # lado a lado
            y_offset = margem
        else:
            x_offset = margem + 90  # centraliza abaixo dos dois
            y_offset = margem + 130

        canvas.create_image(x_offset + 12, y_offset + 17, image=caminhao_img, anchor=NW)

        x_inicial_grade = x_offset + 50
        y_inicial_grade = y_offset + 10

        for i in range(quad_coluna):
            for j in range(quad_linha):
                idx = caminhao_idx * quad_por_caminhao + (quad_coluna - 1 - i) * quad_linha + j
                x1 = x_inicial_grade + j * quad_largura
                y1 = y_inicial_grade + i * quad_altura
                x2 = x1 + quad_largura
                y2 = y1 + quad_altura
                cor = "#0070C0" if idx < total_quads else "#D9D9D9"
                canvas.create_rectangle(x1, y1, x2, y2, fill=cor, outline='black')



def calcular_empilhamento_line_haul(df_saturacao, db_empilhamento):
    empilhamento_rows = []

    base_df = df_saturacao[df_saturacao['EMBALAGEM_BASE'] == 1]
    sobre_df = df_saturacao[df_saturacao['EMBALAGEM_SOBREPOSTA'] == 1]

    for _, base_row in base_df.iterrows():
        for _, sobre_row in sobre_df.iterrows():
            if base_row['COD FORNECEDOR'] == sobre_row['COD FORNECEDOR']:
                fornecedor = base_row['COD FORNECEDOR']
                embal_base = base_row['EMBALAGEM']
                embal_sobre = sobre_row['EMBALAGEM']

                empilhamento_match = db_empilhamento[
                    (db_empilhamento['COD FORNECEDOR'] == fornecedor) &
                    (db_empilhamento['MDR BASE'] == embal_base) &
                    (db_empilhamento['MDR SOBREPOSTA'] == embal_sobre)
                ]

                if empilhamento_match.empty:
                    continue

                capacidade_veiculo = base_row['CAPACIDADE']

                total_base = base_row['TOTAL DE CXS']
                total_sobre = sobre_row['TOTAL DE CXS']

                usadas_base = 0
                usadas_sobre = 0

                # Empilha 1 base com 1 sobreposta (não considera EMPILHAMENTO BASE)
                while total_base >= 1 and total_sobre >= 1:
                    total_base -= 1
                    total_sobre -= 1
                    usadas_base += 1
                    usadas_sobre += 1

                total_empilhado = usadas_base + usadas_sobre
                chave = f"{fornecedor}-{embal_base}-{embal_sobre}"
               
                saturacao = total_empilhado / capacidade_veiculo

                empilhamento_rows.append({
                    'FORNECEDOR': fornecedor,
                    'EMBALAGEM_BASE': embal_base,
                    'EMBALAGEM_SOBREPOSTA': embal_sobre,
                    'CAPACIDADE_VEÍCULO': capacidade_veiculo,
                    'TOTAL_DE_EMBALAGENS_BASE': base_row['TOTAL DE CXS'],
                    'TOTAL_DE_EMBALAGENS_SOBREPOSTA': sobre_row['TOTAL DE CXS'],
                    'TOTAL_DE_EMBALAGENS_BASE_PARA_COMBINAR': usadas_base,
                    'TOTAL_DE_EMBALAGENS_SOBREPOSTA_PARA_COMBINAR': usadas_sobre,
                    'EMBALAGENS_BASE_RESTANTE': total_base,
                    'EMBALAGENS_SOBREPOSTA_RESTANTE': total_sobre,
                    'CHAVE': chave,
                    'TOTAL_EMBALAGENS_EMPILHADAS': total_empilhado,
                    'SATURAÇÃO': saturacao,
                    'EMPILHAMENTO BASE': 1  # fixo, pois é 1:1
                })

    return pd.DataFrame(empilhamento_rows)


def calcular_empilhamento(df_saturacao, db_empilhamento):
    empilhamento_rows = []

    base_df = df_saturacao[df_saturacao['EMBALAGEM_BASE'] == 1]
    sobre_df = df_saturacao[df_saturacao['EMBALAGEM_SOBREPOSTA'] == 1]

    for _, base_row in base_df.iterrows():
        for _, sobre_row in sobre_df.iterrows():
            if base_row['COD FORNECEDOR'] == sobre_row['COD FORNECEDOR']:
                fornecedor = base_row['COD FORNECEDOR']
                embal_base = base_row['EMBALAGEM']
                embal_sobre = sobre_row['EMBALAGEM']

                empilhamento_match = db_empilhamento[
                    (db_empilhamento['COD FORNECEDOR'] == fornecedor) &
                    (db_empilhamento['MDR BASE'] == embal_base) &
                    (db_empilhamento['MDR SOBREPOSTA'] == embal_sobre)
                ]

                if empilhamento_match.empty:
                    continue

                emp_base = empilhamento_match.iloc[0]['EMPILHAMENTO BASE']
                capacidade_veiculo = base_row['CAPACIDADE']

                total_base = base_row['TOTAL DE CXS']
                total_sobre = sobre_row['TOTAL DE CXS']

                usadas_base = 0
                usadas_sobre = 0

                while total_base >= emp_base and total_sobre >= 1:
                    total_base -= emp_base
                    total_sobre -= 1
                    usadas_base += emp_base
                    usadas_sobre += 1

                total_empilhado = usadas_base + usadas_sobre
                chave = f"{fornecedor}-{embal_base}-{embal_sobre}"
                
                saturacao = total_empilhado / capacidade_veiculo

                empilhamento_rows.append({
                    'FORNECEDOR': fornecedor,
                    'EMBALAGEM_BASE': embal_base,
                    'EMBALAGEM_SOBREPOSTA': embal_sobre,
                    'CAPACIDADE_VEÍCULO': capacidade_veiculo,
                    'TOTAL_DE_EMBALAGENS_BASE': base_row['TOTAL DE CXS'],
                    'TOTAL_DE_EMBALAGENS_SOBREPOSTA': sobre_row['TOTAL DE CXS'],
                    'TOTAL_DE_EMBALAGENS_BASE_PARA_COMBINAR': usadas_base,
                    'TOTAL_DE_EMBALAGENS_SOBREPOSTA_PARA_COMBINAR': usadas_sobre,
                    'EMBALAGENS_BASE_RESTANTE': total_base,
                    'EMBALAGENS_SOBREPOSTA_RESTANTE': total_sobre,
                    'CHAVE': chave,
                    'TOTAL_EMBALAGENS_EMPILHADAS': total_empilhado,
                    'SATURAÇÃO': saturacao,
                    'EMPILHAMENTO BASE': emp_base
                })

    return pd.DataFrame(empilhamento_rows)




















def completar_informacoes(tree, veiculo, tree_resumo, canvas_caminhoes, caminhao_img, usar_manual=False,caminho_BD = 'BD'):


    def split_key_logic(code):
        """
        Splits a code by '/'. 
        Returns the second element if a split occurs, otherwise returns the original code.
        """
        # Convert to string just in case, then split
        
        parts = str(code).split('/')
        
        if len(parts) > 1:
            # If the split created more than one part, take the second one (index 1)
            return parts[1].strip() 
            
        else:
            # Otherwise, take the original part (index 0)
            return parts[0].strip()
    try:


        # --- Leitura dos arquivos ---
        template = pd.read_excel('Template.xlsx', dtype={'COD FORNECEDOR': int, 'DESENHO': str})
        template = template[template['QTDE'] > 0]
        
        
        BD_PN = os.path.join(caminho_base,caminho_BD,"BD_CADASTRO_PN.xlsx")
        BD_MDR = os.path.join(caminho_base,caminho_BD,"BD_CADASTRO_MDR.xlsx")
        VEÍCULOS = os.path.join(caminho_base,caminho_BD,"VEÍCULOS.xlsx")
        db_empilhamento = os.path.join(caminho_base,caminho_BD,"BD_EMPILHAMENTO_EMBALAGENS.xlsx")
        db_efi = os.path.join(caminho_base,caminho_BD,"BD_CADASTRO_MDR_PERDA_COMPRIMENTO.xlsx")
        db_efi = os.path.join(caminho_base,caminho_BD,"BD_CADASTRO_MDR_PERDA_COMPRIMENTO.xlsx")
       
        # ------------------Working in the DB structrue------------------
        db_PN = pd.read_excel(BD_PN, sheet_name='BD', dtype={'CÓD. FORNECEDOR': int, 'DESENHO': str})
        db_PN = db_PN.rename(columns={'CÓD. FORNECEDOR': 'COD FORNECEDOR'})

        db_MDR = pd.read_excel(BD_MDR, sheet_name='BD')
        db_MDR = db_MDR.rename(columns={'DESCRIÇÃO2': 'DESCRIÇÃO'})

        db_veiculos = pd.read_excel(VEÍCULOS, sheet_name='VEÍCULOS')

        db_empilhamento = pd.read_excel(db_empilhamento, sheet_name='BD')
        db_empilhamento = db_empilhamento.rename(columns={'CÓD. FORNECEDOR': 'COD FORNECEDOR'})

        db_efi = pd.read_excel(db_efi,sheet_name='BD')

        # --- Normalização de tipos ---
        db_PN['DESENHO ATUALIZAÇÃO'] = pd.to_datetime(db_PN['DESENHO ATUALIZAÇÃO'], errors='coerce')
        db_MDR['VOLUME'] = pd.to_numeric(db_MDR['VOLUME'], errors='coerce')
        db_MDR['MDR PESO'] = pd.to_numeric(db_MDR['MDR PESO'], errors='coerce')
        db_PN['PESO (Kg) MATERIAL'] = pd.to_numeric(db_PN['PESO (Kg) MATERIAL'], errors='coerce')

        db_PN = db_PN.sort_values('DESENHO ATUALIZAÇÃO', ascending=False)

        # Criar chave composta DESENHO+MDR em db_PN
        db_PN['KEY'] = db_PN['DESENHO'].astype(str) + '_' + db_PN['MDR'].astype(str)

        # --- Mapeamentos únicos para .map() seguros ---
        mapa_fornecedores = db_PN.drop_duplicates('COD FORNECEDOR').set_index('COD FORNECEDOR')['FORNECEDOR']

        # Mapas baseados na chave composta
        mapa_pn = db_PN.drop_duplicates('KEY').set_index('KEY')['DESCRIÇÃO']
        mapa_mdr = db_PN.drop_duplicates('KEY').set_index('KEY')['MDR']
        mapa_qme = db_PN.drop_duplicates('KEY').set_index('KEY')['QME']
        mapa_peso_pn = db_PN.drop_duplicates('KEY').set_index('KEY')['PESO (Kg) MATERIAL']

        # Mapas vindos do db_MDR
        mapa_descricao_mdr = db_MDR.drop_duplicates('MDR').set_index('MDR')['DESCRIÇÃO']
        mapa_volume = db_MDR.drop_duplicates('MDR').set_index('MDR')['VOLUME']
        mapa_peso_mdr = db_MDR.drop_duplicates('MDR').set_index('MDR')['MDR PESO']
        mapa_peso_max = db_veiculos.set_index('COD VEICULO')['PESO MAXIMO']

        # --- Enriquecimento do template ---

        # Passo 1: primeiro trazer MDR pelo DESENHO, para podermos montar a KEY
        template['MDR'] = template['DESENHO'].map(
            db_PN.drop_duplicates('DESENHO').set_index('DESENHO')['MDR']
        )


        # Passo 2: agora que já temos MDR no template, podemos montar a KEY
        template['KEY'] = template['DESENHO'].astype(str) + '_' + template['MDR'].astype(str)
        

        # Passo 3: enriquecer com os mapas
        template['PESO_MAXIMO'] = template['VEICULO'].map(mapa_peso_max)
        template['MAP_KEY'] = (template['COD IMS'].fillna(template['COD FORNECEDOR']).astype(str).str.split('/').str[0] )

       
        template['MAP_KEY'] = pd.to_numeric(template['MAP_KEY'], errors='coerce')
        template['FORNECEDOR'] =template['MAP_KEY'].map(mapa_fornecedores)
       
        template = template.drop(columns=['MAP_KEY'])
       
        template['DESCRIÇÃO MATERIAL'] = template['KEY'].map(mapa_pn)
        template['MDR'] = template['KEY'].map(mapa_mdr)  # reforça MDR correto do KEY
        template['DESCRIÇÃO DA EMBALAGEM'] = template['MDR'].map(mapa_descricao_mdr)
        template['QME'] = template['KEY'].map(mapa_qme)

        template['QTD EMBALAGENS'] = np.ceil(template['QTDE'] / template['QME'])

        template['M³'] = round(template['QTD EMBALAGENS'] * template['MDR'].map(mapa_volume), 1)
        template['PESO MAT'] = round(template['QTDE'] * template['KEY'].map(mapa_peso_pn), 1)
        template['PESO MDR'] = round(template['QTD EMBALAGENS'] * template['MDR'].map(mapa_peso_mdr), 1)
        template['PESO TOTAL'] = template['PESO MAT'] + template['PESO MDR']


        template = template[['COD FORNECEDOR', 'FORNECEDOR', 'COD DESTINO', 'DESENHO', 'QTDE', 'DESCRIÇÃO MATERIAL',
                             'MDR', 'DESCRIÇÃO DA EMBALAGEM', 'QME', 'QTD EMBALAGENS', 'TIPO SATURACAO',
                             'VEICULO', 'M³', 'PESO MAT', 'PESO MDR', 'PESO TOTAL', 'PESO_MAXIMO']]

        # --- Construção da aba Saturação ---
        df_saturacao = (
            template.groupby(['COD FORNECEDOR', 'FORNECEDOR', 'MDR'], as_index=False)['QTD EMBALAGENS']
            .sum()
            .rename(columns={'MDR': 'EMBALAGEM', 'QTD EMBALAGENS': 'TOTAL DE CXS'})
        )

        # Recupera a coluna VEICULO para cada fornecedor + embalagem
        col_veiculo = template[['COD FORNECEDOR', 'MDR', 'VEICULO']].drop_duplicates()
        col_veiculo = col_veiculo.rename(columns={'MDR': 'EMBALAGEM'})

        df_saturacao = df_saturacao.merge(col_veiculo, on=['COD FORNECEDOR', 'EMBALAGEM'], how='left')

        mapa_paletizavel = db_MDR.drop_duplicates('MDR').set_index('MDR')['CAIXA PLÁSTICA']
        mapa_cxs_por_pallet = db_MDR.drop_duplicates('MDR').set_index('MDR')['CAIXAS POR PALLET']

        df_saturacao['CX_PALETIZÁVEL'] = df_saturacao['EMBALAGEM'].map(mapa_paletizavel).fillna(0).astype(int)
        df_saturacao['CXS_POR_PALLET'] = df_saturacao.apply(
            lambda row: 1 if row['CX_PALETIZÁVEL'] != 1 else (
                mapa_cxs_por_pallet.get(row['EMBALAGEM'], 1) or 1), axis=1
        )
        df_saturacao['CXS/PALLETS_TOTAL'] = df_saturacao['TOTAL DE CXS'] / df_saturacao['CXS_POR_PALLET']

        valor_veiculo = db_veiculos.loc[db_veiculos['COD VEICULO'] == veiculo, 'VEICULOS'].iloc[0]
        # Mapeia de código do veículo (ex: 4) → coluna de capacidade no db_MDR (ex: "14 x 2,4 x 2,78")
        mapa_coluna_capacidade = db_veiculos.set_index('COD VEICULO')['VEICULOS'].to_dict()
        

        # Garante que os MDRs na base estejam em caixa alta
        db_MDR['MDR'] = db_MDR['MDR'].astype(str).str.upper()

        
        def obter_veiculo_anterior(cod_veic):
            if cod_veic in [4, 5, 6, 7, 8, 9, 14]:
                return 3
            elif cod_veic in [2, 3, 12, 13, 15, 16, 17, 18]:
                return 1
            elif cod_veic == 1:
                return 10
            elif cod_veic == 10:
                return 11
            elif cod_veic == 11:
                return 11
            return None

        def obter_capacidade_por_linha(row):
            mdr = str(row['EMBALAGEM']).upper()  # Converte para string e caixa alta
            cod_veic = row['VEICULO']
            coluna = mapa_coluna_capacidade.get(cod_veic)

            if not coluna:
                # print(f"[ERRO] Código de veículo {cod_veic} não mapeado.")
                return None
            if coluna not in db_MDR.columns:
                # print(f"[ERRO] Coluna '{coluna}' não encontrada no db_MDR para veículo {cod_veic}")
                return None

            filtro = db_MDR['MDR'] == mdr
            capacidade_series = db_MDR.loc[filtro, coluna].dropna()

            if capacidade_series.empty:
                # print(f"[ERRO] Capacidade não encontrada para MDR {mdr} na coluna '{coluna}' (cod veic {cod_veic})")
                return None

            return capacidade_series.values[0]

        def obter_capacidade_por_linha_veic_anterior(row):

            mdr = str(row['EMBALAGEM']).upper()
            cod_veic = int(row['VEICULO'])
            veic_anterior = obter_veiculo_anterior(cod_veic)
           
            if veic_anterior is None :
                # print(f"[INFO] Veículo anterior não definido para código {cod_veic}")
                return None

            coluna = mapa_coluna_capacidade.get(veic_anterior)

            if not coluna:
                # print(f"[ERRO] Código de veículo anterior {veic_anterior} não mapeado.")
                return None
            
            if coluna not in db_MDR.columns:
                # print(f"[ERRO] Coluna '{coluna}' não encontrada no db_MDR para veículo anterior {veic_anterior}")
                return None

            filtro = db_MDR['MDR'].str.contains(mdr)
            capacidade_series = db_MDR.loc[filtro, coluna].dropna()

            if capacidade_series.empty:
                print(
                    f"[ERRO] Capacidade não encontrada para MDR {mdr} na coluna '{coluna}' (veic anterior {veic_anterior})")
                return None

            return capacidade_series.values[0]

        df_saturacao['CAPACIDADE'] = df_saturacao.apply(obter_capacidade_por_linha, axis=1)
        df_saturacao['VEICULO'] = df_saturacao['VEICULO'].fillna(0)
        df_saturacao['VEICULO'] = df_saturacao['VEICULO'].astype(int)         
        df_saturacao['CAPACIDADE_VEIC_ANTERIOR'] = df_saturacao.apply(obter_capacidade_por_linha_veic_anterior, axis=1)

        

        df_saturacao['SATURAÇÃO COM VEÍCULO MENOR (%)'] = round(
            df_saturacao['CXS/PALLETS_TOTAL'] / df_saturacao['CAPACIDADE_VEIC_ANTERIOR'] * 100, 2
        )

        bases = set(zip(db_empilhamento['FORNECEDOR'], db_empilhamento['MDR BASE']))
        sobrepostas = set(zip(db_empilhamento['FORNECEDOR'], db_empilhamento['MDR SOBREPOSTA']))
        df_saturacao['EMBALAGEM_BASE'] = df_saturacao.apply(
            lambda row: 1 if (row['FORNECEDOR'], row['EMBALAGEM']) in bases else 0, axis=1)
        df_saturacao['EMBALAGEM_SOBREPOSTA'] = df_saturacao.apply(
            lambda row: 1 if (row['FORNECEDOR'], row['EMBALAGEM']) in sobrepostas else 0, axis=1)

        df_saturacao['CHAVE'] = df_saturacao['COD FORNECEDOR'].astype(str) + '-' + df_saturacao['EMBALAGEM'].astype(str)


        # --- Eficiência de empilhamento por embalagem (evita .map com índice duplicado) ---
        mapa_efi = db_efi.drop_duplicates('CHAVE FORNE + MDR').set_index('CHAVE FORNE + MDR')[valor_veiculo]
        df_saturacao['EFICIÊNCIA_COMPRIMENTO'] = df_saturacao['CHAVE'].map(mapa_efi).fillna(1)

        mapa_volume_efi = db_MDR.drop_duplicates('CHAVE EMBALAGENS').set_index('CHAVE EMBALAGENS')['VOLUME']
        df_saturacao['M³ POR EMBALAGEM'] = df_saturacao['CHAVE'].map(mapa_volume_efi) * \
                                            df_saturacao['CXS_POR_PALLET'] * df_saturacao['CXS/PALLETS_TOTAL']

        # --- Cálculo de empilhamento ---
        df_calculo_empilhamento = calcular_empilhamento(df_saturacao, db_empilhamento)

        # --- Saturação final por embalagem ---
        def integrar_saturacao_total(df_sat, df_emp):
            def calcular(row):
                filtro = (df_emp['FORNECEDOR'] == row['COD FORNECEDOR']) & \
                         (df_emp['EMBALAGEM_BASE'] == row['EMBALAGEM'])
                soma_saturacoes = df_emp[filtro]['SATURAÇÃO'].sum()
                proporcao = row['CXS/PALLETS_TOTAL'] / row['CAPACIDADE']
                return (proporcao + soma_saturacoes) * row['EFICIÊNCIA_COMPRIMENTO']

            df_sat['SATURAÇÃO_TOTAL'] = df_sat.apply(calcular, axis=1)
            df_sat['SATURAÇÃO_POR_MDR'] = df_sat['SATURAÇÃO_TOTAL'] / df_sat['TOTAL DE CXS']
            return df_sat

        if not df_calculo_empilhamento.empty:
            df_saturacao = integrar_saturacao_total(df_saturacao, df_calculo_empilhamento)
        else:
            df_saturacao['SATURAÇÃO_TOTAL'] = df_saturacao.apply(
                lambda row: row['CXS/PALLETS_TOTAL'] / row['CAPACIDADE'], axis=1)
            df_saturacao['SATURAÇÃO_POR_MDR'] = df_saturacao['SATURAÇÃO_TOTAL'] / df_saturacao['TOTAL DE CXS']

        # --- Cálculo da SAT por linha ---
        template.loc[:, 'CHAVE'] = template['COD FORNECEDOR'].astype(str) + '-' + template['MDR'].astype(str)
        template = template.merge(df_saturacao[['CHAVE', 'SATURAÇÃO_POR_MDR']], on='CHAVE', how='left')
        template['SAT VOLUME (%)'] = round(template['QTD EMBALAGENS'] * template['SATURAÇÃO_POR_MDR'] * 100, 2)
        template['SAT PESO (%)'] = round(template['PESO TOTAL'] / template['PESO_MAXIMO'] * 100, 2)
        
        # --- Capacidade Útil Calculations (A-D) ---
        # C) Combined - limiting factor per row
        template['CAPACIDADE ÚTIL (%)'] = template[['SAT VOLUME (%)', 'SAT PESO (%)']].max(axis=1)
        
        template.drop(columns=['CHAVE', 'SATURAÇÃO_POR_MDR'], inplace=True)
        df_saturacao.drop(columns=['CHAVE'], inplace=True)

        # --- Criação das variáveis para a tabela final ---

        ocupacao = template['SAT VOLUME (%)'].sum()
        qtd_veiculos = ceil(ocupacao / 100)
        volume = template['M³'].sum()
        peso = template['PESO TOTAL'].sum()
        embalagens = template['QTD EMBALAGENS'].sum()
        
        # Get vehicle capacity from db_veiculos
        capacidade_veiculo_m3 = None
        capacidade_veiculo_kg = None
        if not template.empty and 'PESO_MAXIMO' in template.columns:
            capacidade_veiculo_kg = template['PESO_MAXIMO'].iloc[0] if template['PESO_MAXIMO'].notna().any() else None
        
        # Get volume capacity from db_veiculos
        try:
            veiculo_info = db_veiculos[db_veiculos['COD VEICULO'] == veiculo]
            if not veiculo_info.empty and 'CAPACIDADE M³' in veiculo_info.columns:
                capacidade_veiculo_m3 = veiculo_info['CAPACIDADE M³'].iloc[0]
        except:
            pass
        
        # --- Capacidade Útil Calculations for Summary ---
        # A) Volume-based capacity per vehicle
        cap_util_volume_percent = (ocupacao / qtd_veiculos) if qtd_veiculos > 0 else 0
        cap_util_volume_m3 = (volume / qtd_veiculos) if qtd_veiculos > 0 else 0
        
        # D) Remaining capacity
        volume_restante = (capacidade_veiculo_m3 * qtd_veiculos - volume) if capacidade_veiculo_m3 else None

        # Preenche a tree_resumo (que deve ser passada como argumento)
        resumo_dados = [
            ("Ocupação Total", f"{ocupacao:.2f}%"),
            ("Qtd Veículos", qtd_veiculos),
            ("Volume Total", f"{volume:.1f} m³"),
            ("Peso Total", f"{peso:.1f} kg"),
            ("Embalagens", int(embalagens)),
            ("Cap. Útil (m³)", f"{cap_util_volume_m3:.1f} m³"),
            ("Cap. Útil (%)", f"{cap_util_volume_percent:.2f}%"),
        ]
        
        # Add remaining capacity if available
        if volume_restante is not None and volume_restante >= 0:
            resumo_dados.append(("Volume Restante", f"{volume_restante:.1f} m³"))

        linhas_validas = template[
            (template['DESENHO'].notna()) &
            (template['COD FORNECEDOR'].notna()) &
            (template['QTDE'] > 0)
            ].shape[0]

        linha_qme = template[
            (template['QME'] > 0) &
            (template['QTDE'] > 0)
            ].shape[0]

        # Limpa e atualiza a tabela tree_resumo
        tree_resumo.delete(*tree_resumo.get_children())
        for item in resumo_dados:
            tree_resumo.insert("", END, values=item)


        # --- Atualiza TreeView (Tkinter) ---
        tree.delete(*tree.get_children())
        tree["columns"] = list(template.columns)
        tree["show"] = "headings"
        
        # Define reasonable initial widths for each column
        column_widths = {
            'COD FORNECEDOR': 140,
            'FORNECEDOR': 280,
            'COD DESTINO': 140,
            'DESENHO': 150,
            'QTDE': 90,
            'DESCRIÇÃO MATERIAL': 350,
            'MDR': 120,
            'DESCRIÇÃO DA EMBALAGEM': 280,
            'QME': 90,
            'QTD EMBALAGENS': 140,
            'TIPO SATURACAO': 140,
            'VEICULO': 120,
            'M³': 90,
            'PESO MAT': 110,
            'PESO MDR': 110,
            'PESO TOTAL': 120,
            'PESO_MAXIMO': 140,
            'SAT VOLUME (%)': 140,
            'SAT PESO (%)': 140,
            'CAPACIDADE ÚTIL (%)': 150
        }
        
        for col in template.columns:
            tree.heading(col, text=col)
            width = column_widths.get(col, 150)  # Default to 150 if not specified
            tree.column(col, width=width, anchor="center", stretch=True, minwidth=80)
        
        for _, row in template.iterrows():
            tree.insert("", END, values=list(row))

        desenhar_caminhoes(canvas_caminhoes, ocupacao, caminhao_img)

        # --- Exporta para Excel formatado ---
        with pd.ExcelWriter('VIAJANTE.xlsx', engine='openpyxl') as writer:
            template.to_excel(writer, sheet_name='Template Completo', index=False)
            df_saturacao.to_excel(writer, sheet_name='Saturação', index=False)
            df_calculo_empilhamento.to_excel(writer, sheet_name='Calculo Empilhamento', index=False)

            header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            header_font = Font(bold=True, color='000000')
            header_align = Alignment(horizontal='center', vertical='center')

            for sheet_name in ['Template Completo', 'Saturação', 'Calculo Empilhamento']:
                ws = writer.sheets[sheet_name]
                for col_num, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
                    largura = max(len(str(cell.value) or '') for cell in col) + 2
                    ws.column_dimensions[get_column_letter(col_num)].width = largura
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align

            if 'MDR' in template.columns:
                pn_nao_cadastrados = template[
                    template['MDR'].isna() | (template['MDR'].astype(str).str.strip() == '')
                ].copy()

                # select only the requested columns if they exist in the dataframe
                cols_to_keep = ['COD FORNECEDOR', 'FORNECEDOR', 'COD DESTINO', 'DESENHO']
                existing_cols = [c for c in cols_to_keep if c in pn_nao_cadastrados.columns]

                if not pn_nao_cadastrados.empty and existing_cols:
                    pn_nao_cadastrados = pn_nao_cadastrados[existing_cols]
                    pn_nao_cadastrados.drop_duplicates(subset=["DESENHO"], inplace=True)
                    pn_nao_cadastrados.to_excel(writer, sheet_name='PN Não Cadastrados', index=False)

    except Exception as e:

        print(f"Erro: {e}")

        traceback.print_exc()


def consolidar_dados():
    # Carrega os dados
    fluxos_path = os.path.join(caminho_base, "BD", "FLUXO.xlsx")
    fluxos = pd.read_excel(fluxos_path, sheet_name='FLUXOS')
    template = pd.read_excel('VIAJANTE.xlsx', sheet_name='Template Completo')

    # Filtra linhas com quantidade válida e prepara as colunas
    template = template[template['QTDE'] > 0].copy() # Use .copy() to avoid SettingWithCopyWarning
    template['COD FORNECEDOR'] = template['COD FORNECEDOR'].astype(str)


    # --- FIX: Ensure the supplier name column is also a string ---
    template['FORNECEDOR'] = template['FORNECEDOR'].fillna('').astype(str)


    def normalizar_codigos(campo):
        if pd.isna(campo):
            return []
        return re.split(r'\s*/\s*', str(campo).strip())

    dados_volume = []

    for cod_dest in template['COD DESTINO'].dropna().unique():
        subset_template = template[template['COD DESTINO'] == cod_dest]

        fornecedores_template_set = set()
        for cod in subset_template['COD FORNECEDOR'].astype(str):
            fornecedores_template_set.update(normalizar_codigos(cod))

        rotas_destino = fluxos[fluxos['COD DESTINO'].astype(str).str.contains(str(cod_dest))]

       
        
        for _, rota in rotas_destino.iterrows():
            cod_fluxo = rota['COD FLUXO']
            destino = rota['NOME DESTINO']
            veiculo = rota['VEICULO PRINCIPAL']
            tipo_saturacao = rota['TIPO SATURACAO']
            transportadora = rota['TRANSPORTADORA']
            fornecedores_rota = normalizar_codigos(rota['COD FORNECEDOR'])

            fornecedores_comuns = [f for f in fornecedores_rota if f in fornecedores_template_set]

            if fornecedores_comuns:
                linhas_rota = subset_template[subset_template['COD FORNECEDOR'].astype(str).isin(fornecedores_comuns)]

                volume_total = linhas_rota['M³'].sum()
                peso_total = linhas_rota['PESO TOTAL'].sum()
                embalagens_total = linhas_rota['QTD EMBALAGENS'].sum()

                if tipo_saturacao.upper() == 'VOLUME':
                    saturacao_total = linhas_rota['SAT VOLUME (%)'].sum()
                else:
                    saturacao_total = linhas_rota['SAT PESO (%)'].sum()

                nomes_fornecedores = linhas_rota[['COD FORNECEDOR', 'FORNECEDOR']].drop_duplicates()
                nomes_fornecedores['COD FORNECEDOR'] = nomes_fornecedores['COD FORNECEDOR'].astype(str)

                nomes_ordenados = nomes_fornecedores.set_index('COD FORNECEDOR').loc[fornecedores_comuns]['FORNECEDOR'].tolist()


              


                cargas = ceil(saturacao_total / 100) if saturacao_total > 0 else 0

                # --- Coluna de Sugestão ---
                saturacao_residual = saturacao_total % 100
                if cargas > 0 and saturacao_residual <= 2:
                    sugestao = "Cortar coleta do último veículo"
                elif cargas > 0 and saturacao_residual <= 50:
                    sugestao = "Alterar último veículo para menor porte"
                else:
                    sugestao = "Manter coleta"

                # --- Apuração de MDR ---
                coluna_sat = 'SAT VOLUME (%)' if tipo_saturacao.upper() == 'VOLUME' else 'SAT PESO (%)'
                
                linhas_template_todas = subset_template[
                    subset_template['COD FORNECEDOR'].astype(str).isin(fornecedores_comuns)]

                total_desenhos = linhas_template_todas['DESENHO'].nunique()
                desenhos_apurados = linhas_template_todas[
                    linhas_template_todas[coluna_sat].fillna(0) > 0
                    ]['DESENHO'].nunique()

                perc_mdr = round((desenhos_apurados / total_desenhos) * 100, 1) if total_desenhos else 0.0

                # --- Capacidade Útil per route ---
                cap_util_volume_rota_m3 = (volume_total / cargas) if cargas > 0 else 0
                cap_util_volume_rota_percent = (saturacao_total / cargas) if cargas > 0 else 0

                dados_volume.append({
                    'COD FLUXO': cod_fluxo,
                    'COD DESTINO': cod_dest,
                    'DESTINO': destino,
                    'CÓDIGOS FORNECEDORES': ', '.join(fornecedores_comuns),
                    'FORNECEDORES NA ROTA': ', '.join(nomes_ordenados),
                    'VEÍCULO': veiculo,
                    'TECNOLOGIA': rota['TECNOLOGIA'],
                    'MOT': rota['MOT'],
                    'TRANSPORTADORA': transportadora,
                    'TIPO DE SATURAÇÃO': tipo_saturacao,
                    'VOLUME TOTAL (m³)': round(volume_total, 1),
                    'PESO TOTAL (kg)': round(peso_total, 1),
                    'EMBALAGENS TOTAL': int(embalagens_total),
                    'SATURAÇÃO TOTAL (%)': round(saturacao_total, 2),
                    'CARGAS': cargas,
                    'CAP. ÚTIL (m³)': round(cap_util_volume_rota_m3, 1),
                    'CAP. ÚTIL (%)': round(cap_util_volume_rota_percent, 2),
                    'SUGESTÃO': sugestao,
                    '% MDRs APURADOS': perc_mdr
                })

    df_volume = pd.DataFrame(dados_volume)
    df_volume.to_excel('Volume_por_rota.xlsx', index=False)
#tree = ttk.Treeview()
#tree_resumo = ttk.Treeview()
#completar_informacoes(tree,3, tree_resumo)

